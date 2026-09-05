import json
import hashlib
import logging
import unicodedata
import csv
import io
import re
from tempfile import SpooledTemporaryFile
from datetime import datetime, timezone
from decimal import Decimal
from typing import Any

from lxml import html
from sqlalchemy import and_, func, or_, select
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.orm import Session
import requests

from app.models.fiscal_assistant import FiscalSuggestion
from app.core.master_database import MasterSessionLocal
from app.models.master_fiscal_reference import (
    MasterFiscalCollectiveObservation,
    MasterFiscalCollectiveSuggestion,
    MasterIbsCbsClassTrib,
    MasterFiscalCestCode,
    MasterFiscalCfopCode,
    MasterFiscalNcmCode,
    MasterFiscalReferenceSync,
)
from app.models.fiscal import CompanyFiscalSetting
from app.models.product import Product
from app.models.stock_entry import StockEntry, StockEntryItem
from app.schemas.fiscal_assistant import FiscalAlert


logger = logging.getLogger(__name__)


def normalize_fiscal_description(value: str | None) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Za-z0-9]+", " ", text).upper()
    return re.sub(r"\s+", " ", text).strip()


def _clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def _digits(value: str | None) -> str | None:
    if value is None:
        return None
    result = "".join(char for char in value if char.isdigit())
    return result or None


def _is_valid_ncm_code(value: str | None) -> bool:
    """Return true only for an eight-digit NCM item code.

    Four-digit chapters (for example 1006 for rice) are present in some
    official exports and must never be offered as a product NCM.  Padding
    them to ``00001006`` makes the value look like eight digits but SEFAZ
    correctly rejects it.
    """
    code = _digits(value)
    return bool(code and len(code) == 8 and code[:2] != "00")


def _collective_signature(payload: dict[str, Any]) -> str:
    """Stable signature for an aggregate fiscal classification."""
    parts = [
        str(payload.get(field) or "")
        for field in (
            "normalized_description", "barcode", "ncm", "cest", "cfop",
            "origin", "cst", "csosn", "ibs_cbs_cst",
            "ibs_cbs_classification", "selective_tax_cst",
            "selective_tax_classification",
        )
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def _company_fingerprint(db: Session) -> str | None:
    """Return an irreversible company fingerprint; never copy company data."""
    setting = fiscal_context_for_company(db)
    company_id = _digits(setting.cnpj) if setting is not None else None
    if not company_id:
        return None
    return hashlib.sha256(
        f"lyncar-fiscal-collective-v1|{company_id}".encode("utf-8")
    ).hexdigest()


def _learn_collectively(db: Session, payload: dict[str, Any]) -> None:
    """Store anonymous aggregate evidence without affecting another tenant."""
    if not _is_valid_ncm_code(payload.get("ncm")):
        return
    fingerprint = _company_fingerprint(db)
    if fingerprint is None:
        return
    signature = _collective_signature(payload)
    values = {
        field: payload.get(field)
        for field in (
            "normalized_description", "barcode", "unit", "ncm", "cest", "cfop",
            "origin", "cst", "csosn", "ibs_cbs_cst",
            "ibs_cbs_classification", "selective_tax_cst",
            "selective_tax_classification",
        )
    }
    try:
        with MasterSessionLocal() as master_db:
            aggregate = master_db.scalar(
                select(MasterFiscalCollectiveSuggestion).where(
                    MasterFiscalCollectiveSuggestion.signature == signature
                )
            )
            if aggregate is None:
                aggregate = MasterFiscalCollectiveSuggestion(
                    signature=signature,
                    confirmations_count=1,
                    companies_count=0,
                    **values,
                )
                master_db.add(aggregate)
            else:
                aggregate.confirmations_count = (aggregate.confirmations_count or 0) + 1
                aggregate.last_confirmed_at = datetime.now(timezone.utc)

            observation = master_db.scalar(
                select(MasterFiscalCollectiveObservation).where(
                    MasterFiscalCollectiveObservation.suggestion_signature == signature,
                    MasterFiscalCollectiveObservation.company_fingerprint == fingerprint,
                )
            )
            if observation is None:
                master_db.add(MasterFiscalCollectiveObservation(
                    suggestion_signature=signature,
                    company_fingerprint=fingerprint,
                    confirmations_count=1,
                ))
                aggregate.companies_count = (aggregate.companies_count or 0) + 1
            else:
                observation.confirmations_count = (observation.confirmations_count or 0) + 1
                observation.last_confirmed_at = datetime.now(timezone.utc)
            master_db.commit()
    except Exception:
        # A temporary master outage must not block stock/product operations.
        logger.exception("Falha ao registrar aprendizado fiscal coletivo")


def collective_fiscal_suggestions(
    *, description: str | None, barcode: str | None = None, limit: int = 5
) -> list[MasterFiscalCollectiveSuggestion]:
    """Return only cross-company aggregates, never tenant data or auto-fill."""
    search_description = normalize_fiscal_description(description)
    search_barcode = _clean(barcode)
    clauses = []
    if search_barcode:
        clauses.append(MasterFiscalCollectiveSuggestion.barcode == search_barcode)
    if search_description:
        words = [word for word in search_description.split() if len(word) >= 3][:5]
        if words:
            clauses.append(and_(*[
                MasterFiscalCollectiveSuggestion.normalized_description.ilike(f"%{word}%")
                for word in words
            ]))
    if not clauses:
        return []
    try:
        with MasterSessionLocal() as master_db:
            return list(master_db.scalars(
                select(MasterFiscalCollectiveSuggestion)
                .where(
                    MasterFiscalCollectiveSuggestion.active.is_(True),
                    MasterFiscalCollectiveSuggestion.companies_count >= 2,
                    MasterFiscalCollectiveSuggestion.confirmations_count >= 2,
                    or_(*clauses),
                )
                .order_by(
                    MasterFiscalCollectiveSuggestion.companies_count.desc(),
                    MasterFiscalCollectiveSuggestion.confirmations_count.desc(),
                    MasterFiscalCollectiveSuggestion.last_confirmed_at.desc(),
                )
                .limit(limit)
            ).all())
    except Exception:
        logger.exception("Falha ao consultar sugestões fiscais coletivas")
        return []


def fiscal_context_for_company(db: Session) -> CompanyFiscalSetting | None:
    return db.scalar(select(CompanyFiscalSetting).order_by(CompanyFiscalSetting.id.asc()))


def _is_simples_context(setting: CompanyFiscalSetting | None) -> bool:
    regime = (_clean(setting.tax_regime) if setting else None) or ""
    crt = (_clean(setting.crt) if setting else None) or ""
    return regime in {"mei", "simples_nacional"} or crt in {"1", "2", "4"}


def _is_mei_context(setting: CompanyFiscalSetting | None) -> bool:
    regime = (_clean(setting.tax_regime) if setting else None) or ""
    crt = (_clean(setting.crt) if setting else None) or ""
    return regime == "mei" or crt == "4"


def default_tax_suggestion_for_product(
    product: Product | None,
    *,
    description: str | None = None,
    barcode: str | None = None,
    fiscal_setting: CompanyFiscalSetting | None = None,
) -> FiscalSuggestion:
    name = description or (product.name if product else None) or "Produto"
    normalized = normalize_fiscal_description(name)
    is_mei = _is_mei_context(fiscal_setting)
    is_simples = _is_simples_context(fiscal_setting)
    default_csosn = "102" if (is_mei or is_simples) else None
    default_cst = None if is_simples else "00"
    source_note = "default_mei" if is_mei else ("default_simples" if is_simples else "default_regime_normal")
    return FiscalSuggestion(
        id=0,
        normalized_description=normalized,
        original_description=name,
        barcode=_clean(barcode or (product.barcode if product else None)),
        unit=_clean(product.unit if product else None),
        ncm=_clean(product.ncm if product else None),
        cest=_clean(product.cest if product else None),
        cfop=_clean(product.cfop_sale if product else None) or "5102",
        origin=_clean(product.origin if product else None) or "0",
        cst=_clean(product.cst if product else None) or default_cst,
        csosn=_clean(product.csosn if product else None) or default_csosn,
        icms_rate=product.icms_rate if product else None,
        pis_rate=product.pis_rate if product else None,
        cofins_rate=product.cofins_rate if product else None,
        ipi_rate=product.ipi_rate if product else None,
        ibs_cbs_cst=_clean(product.ibs_cbs_cst if product else None),
        ibs_cbs_classification=_clean(product.ibs_cbs_classification if product else None),
        cbs_rate=product.cbs_rate if product else None,
        ibs_state_rate=product.ibs_state_rate if product else None,
        ibs_city_rate=product.ibs_city_rate if product else None,
        selective_tax_cst=_clean(product.selective_tax_cst if product else None),
        selective_tax_classification=_clean(product.selective_tax_classification if product else None),
        selective_tax_rate=product.selective_tax_rate if product else None,
        source=source_note,
        source_reference=(
            f"regime={fiscal_setting.tax_regime or '-'};crt={fiscal_setting.crt or '-'}"
            if fiscal_setting is not None
            else "sem perfil fiscal da empresa"
        ),
        usage_count=0,
        last_used_at=datetime.now(timezone.utc),
        notes=(
            "Sugestao padrao do assistente para venda. XML de entrada e historico de compra "
            "nao definem sozinho a tributacao de saida; confirmar com contador."
        ),
    )


def _as_list(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("Nomenclaturas", "nomenclaturas", "data", "items", "results", "registros"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
    return []


def _first_value(item: dict[str, Any], *keys: str) -> Any:
    lowered = {str(key).lower(): value for key, value in item.items()}
    for key in keys:
        if key in item and item[key] not in (None, ""):
            return item[key]
        value = lowered.get(key.lower())
        if value not in (None, ""):
            return value
    return None


def _upsert_sync_status(
    db: Session,
    source_type: str,
    *,
    source_name: str,
    source_url: str | None,
    status: str,
    records_loaded: int,
    message: str | None = None,
) -> MasterFiscalReferenceSync:
    existing = db.scalar(
        select(MasterFiscalReferenceSync).where(MasterFiscalReferenceSync.source_type == source_type)
    )
    if existing is None:
        existing = MasterFiscalReferenceSync(source_type=source_type, source_name=source_name)
        db.add(existing)
    existing.source_name = source_name
    existing.source_url = source_url
    existing.status = status
    existing.records_loaded = records_loaded
    existing.message = message
    existing.synced_at = datetime.now(timezone.utc)
    return existing


def sync_ncm_from_json_payload(db: Session, payload: Any, *, source_url: str | None = None) -> int:
    records = _as_list(payload)
    # The Siscomex download contains the complete NCM tree: chapter, heading,
    # subheading and item.  An item description can be only "Polido ou
    # brunido", while its parent is "Arroz".  Preserve that official path in
    # the search index so commercial names work for *all* sectors, not only
    # for hand-maintained synonym lists.
    official_tree: list[tuple[str, str]] = []
    for item in records:
        raw_code = _digits(str(_first_value(item, "Codigo", "codigo", "co_ncm", "ncm", "code") or ""))
        description = _clean(
            str(_first_value(item, "Descricao", "descricao", "no_ncm_por", "description", "nome") or "")
        )
        if raw_code and description:
            official_tree.append((raw_code, description))

    loaded = 0
    seen_codes: set[str] = set()
    for code, description in official_tree:
        if not code or not description:
            continue
        # Do not import headings/subheadings as product NCMs.  In
        # particular, zfilling a four-digit chapter creates invalid values
        # such as 00001006, which must not reach a fiscal document.
        if len(code) != 8 or not _is_valid_ncm_code(code):
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)
        row = db.scalar(select(MasterFiscalNcmCode).where(MasterFiscalNcmCode.code == code))
        if row is None:
            row = MasterFiscalNcmCode(code=code, description=description, normalized_description="")
            db.add(row)
        row.description = description
        hierarchy = [
            ancestor_description
            for ancestor_code, ancestor_description in official_tree
            if len(ancestor_code) <= len(code) and code.startswith(ancestor_code)
        ]
        row.normalized_description = normalize_fiscal_description(" ".join(hierarchy))
        row.start_date = _clean(str(_first_value(item, "Data_Inicio", "data_inicio", "inicio") or "")) or row.start_date
        row.end_date = _clean(str(_first_value(item, "Data_Fim", "data_fim", "fim") or "")) or row.end_date
        row.active = row.end_date in (None, "", "31/12/9999", "9999-12-31")
        row.source = "siscomex_classif"
        loaded += 1
    _upsert_sync_status(
        db,
        "ncm",
        source_name="NCM oficial - Portal Unico Siscomex/Classif",
        source_url=source_url,
        status="success",
        records_loaded=loaded,
        message="Tabela NCM sincronizada.",
    )
    return loaded


def sync_cfop_from_rows(db: Session, rows: list[dict[str, Any]], *, source_url: str | None = None) -> int:
    loaded = 0
    for item in rows:
        code = _digits(str(_first_value(item, "CFOP", "codigo", "code") or ""))
        description = _clean(str(_first_value(item, "Descricao", "descricao", "description") or ""))
        if not code or not description:
            continue
        row = db.scalar(select(MasterFiscalCfopCode).where(MasterFiscalCfopCode.code == code))
        if row is None:
            row = MasterFiscalCfopCode(code=code, description=description)
            db.add(row)
        row.description = description
        row.direction = "entrada" if code.startswith(("1", "2", "3")) else "saida"
        row.operation_type = "interestadual" if code.startswith(("2", "6")) else "interna"
        if code.startswith(("3", "7")):
            row.operation_type = "exterior"
        row.source = "confaz_sinief"
        row.active = True
        loaded += 1
    _upsert_sync_status(
        db,
        "cfop",
        source_name="CFOP oficial - CONFAZ/Ajustes SINIEF",
        source_url=source_url,
        status="success",
        records_loaded=loaded,
        message="Tabela CFOP sincronizada.",
    )
    return loaded


def sync_cest_from_rows(db: Session, rows: list[dict[str, Any]], *, source_url: str | None = None) -> int:
    loaded = 0
    for item in rows:
        cest = _digits(str(_first_value(item, "CEST", "cest") or ""))
        ncm = _digits(str(_first_value(item, "NCM", "ncm") or ""))
        description = _clean(str(_first_value(item, "Descricao", "descricao", "description") or ""))
        if not cest or not description:
            continue
        row = db.scalar(
            select(MasterFiscalCestCode).where(
                MasterFiscalCestCode.cest == cest,
                MasterFiscalCestCode.ncm == ncm,
                MasterFiscalCestCode.description == description,
            )
        )
        if row is None:
            row = MasterFiscalCestCode(cest=cest, ncm=ncm, description=description)
            db.add(row)
        row.segment = _clean(str(_first_value(item, "Segmento", "segmento", "segment") or "")) or row.segment
        row.source = "confaz_cest"
        row.active = True
        loaded += 1
    _upsert_sync_status(
        db,
        "cest",
        source_name="CEST oficial - CONFAZ Convenios ICMS 92/15 e 142/18",
        source_url=source_url,
        status="success",
        records_loaded=loaded,
        message="Tabela CEST sincronizada.",
    )
    return loaded


def sync_ibs_cbs_class_trib_from_rows(
    db: Session,
    rows: list[dict[str, Any]],
    *,
    source_url: str | None = None,
) -> int:
    loaded = 0
    for item in rows:
        cst = _digits(str(_first_value(item, "CST", "cst", "CST-IBS/CBS", "CST IBS/CBS") or ""))
        cclass = _digits(
            str(
                _first_value(
                    item,
                    "cClassTrib",
                    "cclass_trib",
                    "Classificacao",
                    "Classificação",
                    "codigo_classificacao",
                )
                or ""
            )
        )
        if len(cst) != 3 or len(cclass) != 6:
            continue
        row = db.scalar(select(MasterIbsCbsClassTrib).where(MasterIbsCbsClassTrib.cclass_trib == cclass))
        if row is None:
            row = MasterIbsCbsClassTrib(cst=cst, cclass_trib=cclass)
            db.add(row)
        row.cst = cst
        row.cst_description = _clean(
            str(_first_value(item, "Descricao CST", "Descrição CST", "descricao_cst", "cst_description") or "")
        )
        row.name = _clean(str(_first_value(item, "Nome", "nome", "name") or ""))
        row.description = _clean(
            str(
                _first_value(
                    item,
                    "Descricao",
                    "Descrição",
                    "descricao",
                    "description",
                    "Descricao cClassTrib",
                    "Descrição cClassTrib",
                )
                or ""
            )
        )
        row.group_type = _clean(str(_first_value(item, "Grupo", "grupo", "group_type") or ""))
        row.requires_gibscbs = cst not in {"400", "410"}
        cst_reduction = _first_value(item, "IndReducaoAliq", "ind_reducao_aliq")
        ibs_reduction = _first_value(item, "PercRedIbs", "perc_red_ibs")
        cbs_reduction = _first_value(item, "PercRedCbs", "perc_red_cbs")
        row.ibs_rate_reduction_percent = Decimal(str(ibs_reduction or 0))
        row.cbs_rate_reduction_percent = Decimal(str(cbs_reduction or 0))
        row.requires_rate_reduction = bool(cst_reduction) or bool(
            row.ibs_rate_reduction_percent or row.cbs_rate_reduction_percent
        )
        row.active = True
        loaded += 1
    _upsert_sync_status(
        db,
        "ibs_cbs_class_trib",
        source_name="Tabela cClassTrib IBS/CBS - Portal NF-e",
        source_url=source_url,
        status="success",
        records_loaded=loaded,
        message="Tabela CST/cClassTrib IBS/CBS sincronizada.",
    )
    return loaded


def _rows_from_html_table(text: str) -> list[dict[str, Any]]:
    document = html.fromstring(text)
    rows: list[dict[str, Any]] = []
    for table in document.xpath(".//table"):
        table_rows = table.xpath(".//tr")
        headers: list[str] = []
        for index, tr in enumerate(table_rows):
            cells = [
                re.sub(r"\s+", " ", " ".join(cell.xpath(".//text()"))).strip()
                for cell in tr.xpath("./th|./td")
            ]
            cells = [cell for cell in cells if cell]
            if not cells:
                continue
            if index == 0 or tr.xpath("./th"):
                headers = cells
                continue
            if headers and len(headers) == len(cells):
                rows.append(dict(zip(headers, cells, strict=False)))
            elif len(cells) >= 2:
                rows.append({f"col{position + 1}": value for position, value in enumerate(cells)})
    return rows


def _cfop_rows_from_official_html(text: str) -> list[dict[str, Any]]:
    normalized = html.fromstring(text).text_content()
    normalized = normalized.replace("\xa0", " ")
    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    pattern = re.compile(
        r"(?m)\b([1-7]\.\d{3})\s*[–-]\s*([^\n\r]+?)(?=\n|\r|$)"
    )
    for match in pattern.finditer(normalized):
        code = match.group(1).replace(".", "")
        description = re.sub(r"\s+", " ", match.group(2)).strip(" .;")
        if not code or not description or code in seen:
            continue
        seen.add(code)
        rows.append({"CFOP": code, "Descricao": description})
    return rows


def _cest_rows_from_official_html(text: str) -> list[dict[str, Any]]:
    table_rows = _rows_from_html_table(text)
    rows: list[dict[str, Any]] = []
    for item in table_rows:
        joined = " | ".join(str(value) for value in item.values())
        cest_match = re.search(r"\b(\d{2}\.?\d{3}\.?\d{2})\b", joined)
        if cest_match is None:
            continue
        cest = cest_match.group(1)
        ncm_candidates = [
            match.group(1)
            for match in re.finditer(r"\b(\d{4}(?:\.\d{2})?(?:\.\d{2})?|\d{8})\b", joined)
            if match.group(1) != cest
        ]
        description = ""
        for value in item.values():
            value_text = re.sub(r"\s+", " ", str(value)).strip()
            if not value_text:
                continue
            if cest in value_text or any(candidate in value_text for candidate in ncm_candidates):
                continue
            if len(value_text) > len(description):
                description = value_text
        rows.append(
            {
                "CEST": cest,
                "NCM": ncm_candidates[0] if ncm_candidates else None,
                "Descricao": description or joined,
            }
        )
    if rows:
        return rows

    normalized = html.fromstring(text).text_content().replace("\xa0", " ")
    pattern = re.compile(
        r"\b(\d{2}\.?\d{3}\.?\d{2})\b\s+((?:\d{4}(?:\.\d{2})?(?:\.\d{2})?|\d{8})(?:\s*,\s*)?)*\s*([^\n\r]+)"
    )
    for match in pattern.finditer(normalized):
        rows.append(
            {
                "CEST": match.group(1),
                "NCM": _digits(match.group(2) or ""),
                "Descricao": re.sub(r"\s+", " ", match.group(3)).strip(),
            }
        )
    return rows


def _ibs_cbs_rows_from_official_html(text: str) -> list[dict[str, Any]]:
    rows = _rows_from_html_table(text)
    if rows:
        table_rows = [
            row
            for row in rows
            if _digits(str(_first_value(row, "cClassTrib", "CodClassTrib", "Classificacao") or ""))
        ]
        if table_rows:
            return table_rows

    match = re.search(r"var\s+dadosOriginais\s*=\s*", text)
    if match is None:
        return []

    decoder = json.JSONDecoder()
    try:
        payload, _ = decoder.raw_decode(text[match.end() :])
    except json.JSONDecodeError:
        return []

    parsed_rows: list[dict[str, Any]] = []
    for cst_group in _as_list(payload):
        cst = str(cst_group.get("Cst") or cst_group.get("CST") or "")
        cst_description = str(cst_group.get("NomeCst") or cst_group.get("Descricao CST") or "")
        for item in _as_list(cst_group.get("ClassificacoesTributarias")):
            parsed_rows.append(
                {
                    "CST": item.get("Cst") or cst,
                    "Descricao CST": cst_description,
                    "cClassTrib": item.get("CodClassTrib"),
                    "Nome": item.get("NomeReduzido") or item.get("NomeClassTrib"),
                    "Descricao": item.get("NomeClassTrib") or item.get("NomeReduzido"),
                    "Grupo": item.get("TipoAliq"),
                    "IndReducaoAliq": cst_group.get("IndReducaoAliq"),
                    "PercRedIbs": item.get("PercRedIbs"),
                    "PercRedCbs": item.get("PercRedCbs"),
                }
            )
    return parsed_rows


def _rows_from_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depende do pacote instalado no deploy
        raise RuntimeError(
            "Importacao XLSX exige openpyxl instalado no backend."
        ) from exc

    with SpooledTemporaryFile(max_size=10_000_000) as handle:
        handle.write(content)
        handle.seek(0)
        workbook = load_workbook(handle, read_only=True, data_only=True)
        worksheet = workbook.active
        headers: list[str] = []
        rows: list[dict[str, Any]] = []
        for raw_row in worksheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in raw_row]
            if not any(values):
                continue
            if not headers:
                headers = values
                continue
            if len(values) < len(headers):
                values.extend([""] * (len(headers) - len(values)))
            rows.append(dict(zip(headers, values, strict=False)))
        return rows


def sync_reference_from_url(db: Session, source_type: str, source_url: str) -> int:
    response = requests.get(
        source_url,
        timeout=120,
        headers={
            "User-Agent": "LyncarFiscalAssistant/1.0 (+https://lyncar.com.br)",
            "Accept": "application/json,text/csv,text/html,*/*",
        },
    )
    response.raise_for_status()
    content_type = response.headers.get("content-type", "").lower()
    if source_type == "ncm":
        return sync_ncm_from_json_payload(db, response.json(), source_url=source_url)
    if (
        "spreadsheet" in content_type
        or "excel" in content_type
        or source_url.lower().endswith((".xlsx", ".xlsm"))
    ):
        rows = _rows_from_xlsx(response.content)
    elif "json" in content_type or source_url.lower().endswith(".json"):
        rows = _as_list(response.json())
    else:
        text = response.content.decode(response.encoding or "utf-8-sig", errors="replace")
        if "html" in content_type or "<html" in text[:500].lower():
            if source_type == "cfop":
                rows = _cfop_rows_from_official_html(text)
            elif source_type == "cest":
                rows = _cest_rows_from_official_html(text)
            elif source_type in {"ibs_cbs_class_trib", "cclass_trib", "cclasstrib"}:
                rows = _ibs_cbs_rows_from_official_html(text)
            else:
                rows = _rows_from_html_table(text)
        else:
            rows = list(csv.DictReader(io.StringIO(text), delimiter=";" if ";" in text[:500] else ","))
    if source_type == "cfop":
        return sync_cfop_from_rows(db, rows, source_url=source_url)
    if source_type == "cest":
        return sync_cest_from_rows(db, rows, source_url=source_url)
    if source_type in {"ibs_cbs_class_trib", "cclass_trib", "cclasstrib"}:
        return sync_ibs_cbs_class_trib_from_rows(db, rows, source_url=source_url)
    raise ValueError("Fonte fiscal desconhecida.")


def fiscal_reference_status(db: Session) -> list[MasterFiscalReferenceSync]:
    return list(
        db.scalars(
            select(MasterFiscalReferenceSync).order_by(MasterFiscalReferenceSync.source_type.asc())
        ).all()
    )


def ncm_suggestions(db: Session, description: str | None, *, limit: int = 20) -> list[MasterFiscalNcmCode]:
    """Return official NCM candidates from the complete official hierarchy.

    The user needs to see the alternatives within a family (for example
    common, parboiled and broken rice) instead of the first textual hit being
    mistaken for an automatic classification.
    """
    normalized = normalize_fiscal_description(description)
    # Ignore packaging/quantity tokens such as 5KG, 500ML and 12UN.  They are
    # commercial presentation, not NCM classification terms, and otherwise
    # create accidental matches in unrelated official descriptions.
    words = [
        word
        for word in normalized.split(" ")
        if len(word) >= 3 and len(re.findall(r"[A-Z]", word)) >= 3
    ][:8]
    if not words:
        return []

    base_filter = (
        MasterFiscalNcmCode.active.is_(True),
        ~MasterFiscalNcmCode.code.startswith("00"),
    )
    # Prefer a match in the item's own official description.  The hierarchy
    # has legal exclusions in some chapter text ("except rice", for example),
    # which must not turn an unrelated chapter into a rice candidate.
    direct_rows = list(
        db.scalars(
            select(MasterFiscalNcmCode)
            .where(
                *base_filter,
                or_(*[MasterFiscalNcmCode.description.ilike(f"%{word}%") for word in words]),
            )
            .order_by(MasterFiscalNcmCode.code.asc())
            .limit(limit * 5)
        ).all()
    )
    family_prefixes = {row.code[:4] for row in direct_rows if len(row.code) == 8}
    if family_prefixes:
        rows = list(
            db.scalars(
                select(MasterFiscalNcmCode)
                .where(
                    *base_filter,
                    or_(*[MasterFiscalNcmCode.code.startswith(prefix) for prefix in family_prefixes]),
                )
                .order_by(MasterFiscalNcmCode.code.asc())
                .limit(limit * 5)
            ).all()
        )
    else:
        rows = list(
            db.scalars(
                select(MasterFiscalNcmCode)
                .where(
                    *base_filter,
                    or_(
                        *[
                            MasterFiscalNcmCode.normalized_description.ilike(f"%{word}%")
                            for word in words
                        ],
                    ),
                )
                .order_by(MasterFiscalNcmCode.code.asc())
                .limit(limit * 25)
            ).all()
        )

    sale_product_words = {
        "BISCOITO",
        "BOLACHA",
        "CHOCOLATE",
        "PAO",
        "BOLO",
        "ARROZ",
        "FEIJAO",
        "LEITE",
        "QUEIJO",
        "MUSSARELA",
        "ACUCAR",
        "TRIGO",
        "CARNE",
        "FRANGO",
        "BEBIDA",
        "REFRIGERANTE",
    }
    industrial_words = {
        "MAQUINA",
        "MAQUINAS",
        "APARELHO",
        "APARELHOS",
        "FABRICAR",
        "FABRICACAO",
        "INDUSTRIA",
        "INDUSTRIAS",
        "PRODUCAO",
    }

    def score(row: MasterFiscalNcmCode) -> tuple[int, int, int, int]:
        row_description = normalize_fiscal_description(row.description)
        matches = sum(1 for word in words if word in row_description)
        consecutive_bonus = 0
        for size in range(min(4, len(words)), 1, -1):
            phrase = " ".join(words[:size])
            if phrase in row_description:
                consecutive_bonus = size
                break
        sale_product_bonus = 1 if any(word in sale_product_words for word in words) and row.code.startswith(("02", "03", "04", "07", "08", "09", "10", "11", "12", "15", "16", "17", "18", "19", "20", "21", "22")) else 0
        industrial_penalty = 1 if any(word in row_description for word in industrial_words) and not any(word in industrial_words for word in words) else 0
        return (-matches, -consecutive_bonus, industrial_penalty, -sale_product_bonus)

    rows.sort(key=score)
    return [row for row in rows if _is_valid_ncm_code(row.code)][:limit]


def ibs_cbs_class_trib_suggestions(
    db: Session,
    description: str | None,
    *,
    cst: str | None = None,
    ncm: str | None = None,
    limit: int = 5,
) -> list[MasterIbsCbsClassTrib]:
    # A descrição ajuda a encontrar NCM, mas não prova um tratamento IBS/CBS
    # especial. Sem uma tabela oficial que relacione o NCM à hipótese fiscal,
    # sugerir Cesta Básica por termos como "biscoito" ou "chocolate" induz o
    # usuário a uma redução indevida. Por isso, as opções especiais só são
    # exibidas quando o CST já foi definido de forma explícita no produto.
    clean_cst = _digits(cst)
    if clean_cst and clean_cst != "000":
        return list(
            db.scalars(
                select(MasterIbsCbsClassTrib)
                .where(
                    MasterIbsCbsClassTrib.active.is_(True),
                    MasterIbsCbsClassTrib.cst == clean_cst,
                )
                .order_by(MasterIbsCbsClassTrib.cclass_trib.asc())
                .limit(limit)
            ).all()
        )

    standard = db.scalar(
        select(MasterIbsCbsClassTrib).where(
            MasterIbsCbsClassTrib.active.is_(True),
            MasterIbsCbsClassTrib.cclass_trib == "000001",
        )
    )
    return [standard] if standard is not None else []


def learn_from_stock_entry_item(
    db: Session,
    item: StockEntryItem,
    *,
    entry: StockEntry | None = None,
    source: str = "stock_entry",
) -> None:
    description = _clean(item.description)
    normalized = normalize_fiscal_description(description)
    if not normalized:
        return
    has_fiscal_data = any(
        _clean(getattr(item, field, None)) is not None
        for field in (
            "ncm",
            "cfop",
            "origin",
            "cst",
            "csosn",
            "icms_rate",
            "pis_rate",
            "cofins_rate",
            "ipi_rate",
            "ibs_cbs_cst",
            "ibs_cbs_classification",
            "selective_tax_cst",
            "selective_tax_classification",
        )
    )
    if not has_fiscal_data:
        return
    source_reference = None
    if entry is not None:
        source_reference = entry.invoice_key or entry.invoice_number or str(entry.id)
    payload = {
        "normalized_description": normalized,
        "original_description": description,
        "barcode": _clean(item.barcode),
        "unit": _clean(item.unit or item.invoice_unit),
        "ncm": _clean(item.ncm),
        "cest": None,
        "cfop": _clean(item.cfop),
        "origin": _clean(item.origin),
        "cst": _clean(item.cst),
        "csosn": _clean(item.csosn),
        "icms_rate": item.icms_rate,
        "pis_rate": item.pis_rate,
        "cofins_rate": item.cofins_rate,
        "ipi_rate": item.ipi_rate,
        "ibs_cbs_cst": _clean(item.ibs_cbs_cst),
        "ibs_cbs_classification": _clean(item.ibs_cbs_classification),
        "cbs_rate": item.cbs_rate,
        "ibs_state_rate": item.ibs_state_rate,
        "ibs_city_rate": item.ibs_city_rate,
        "selective_tax_cst": _clean(item.selective_tax_cst),
        "selective_tax_classification": _clean(item.selective_tax_classification),
        "selective_tax_rate": item.selective_tax_rate,
        "source": source,
        "source_reference": source_reference,
    }
    statement = insert(FiscalSuggestion).values(**payload)
    statement = statement.on_conflict_do_update(
        constraint="uq_fiscal_suggestion_signature",
        set_={
            "original_description": payload["original_description"],
            "unit": payload["unit"],
            "source": payload["source"],
            "source_reference": payload["source_reference"],
            "usage_count": FiscalSuggestion.usage_count + 1,
            "last_used_at": func.now(),
        },
    )
    db.execute(statement)
    _learn_collectively(db, payload)


def learn_from_product(db: Session, product: Product, *, source: str = "product") -> None:
    normalized = normalize_fiscal_description(product.name)
    if not normalized:
        return
    has_fiscal_data = any(
        _clean(getattr(product, field, None)) is not None
        for field in (
            "ncm",
            "cest",
            "cfop_sale",
            "origin",
            "cst",
            "csosn",
            "ibs_cbs_cst",
            "ibs_cbs_classification",
            "selective_tax_cst",
            "selective_tax_classification",
        )
    )
    if not has_fiscal_data:
        return
    payload = {
        "normalized_description": normalized,
        "original_description": product.name,
        "barcode": _clean(product.barcode),
        "unit": _clean(product.unit),
        "ncm": _clean(product.ncm),
        "cest": _clean(product.cest),
        "cfop": _clean(product.cfop_sale),
        "origin": _clean(product.origin),
        "cst": _clean(product.cst),
        "csosn": _clean(product.csosn),
        "icms_rate": product.icms_rate,
        "pis_rate": product.pis_rate,
        "cofins_rate": product.cofins_rate,
        "ipi_rate": product.ipi_rate,
        "ibs_cbs_cst": _clean(product.ibs_cbs_cst),
        "ibs_cbs_classification": _clean(product.ibs_cbs_classification),
        "cbs_rate": product.cbs_rate,
        "ibs_state_rate": product.ibs_state_rate,
        "ibs_city_rate": product.ibs_city_rate,
        "selective_tax_cst": _clean(product.selective_tax_cst),
        "selective_tax_classification": _clean(product.selective_tax_classification),
        "selective_tax_rate": product.selective_tax_rate,
        "source": source,
        "source_reference": str(product.id),
    }
    statement = insert(FiscalSuggestion).values(**payload)
    statement = statement.on_conflict_do_update(
        constraint="uq_fiscal_suggestion_signature",
        set_={
            "original_description": payload["original_description"],
            "unit": payload["unit"],
            "source": payload["source"],
            "source_reference": payload["source_reference"],
            "usage_count": FiscalSuggestion.usage_count + 1,
            "last_used_at": func.now(),
        },
    )
    db.execute(statement)
    _learn_collectively(db, payload)


def fiscal_suggestions_for_product(
    db: Session,
    *,
    product: Product | None = None,
    description: str | None = None,
    barcode: str | None = None,
    limit: int = 5,
) -> list[FiscalSuggestion]:
    search_description = normalize_fiscal_description(description or (product.name if product else None))
    search_barcode = _clean(barcode or (product.barcode if product else None))
    clauses = []
    if search_barcode:
        clauses.append(FiscalSuggestion.barcode == search_barcode)
    if search_description:
        words = [word for word in search_description.split(" ") if len(word) >= 3][:5]
        if words:
            clauses.append(
                and_(
                    *[
                        FiscalSuggestion.normalized_description.ilike(f"%{word}%")
                        for word in words
                    ]
                )
            )
    if not clauses:
        return []
    return list(
        db.scalars(
            select(FiscalSuggestion)
            .where(or_(*clauses))
            .order_by(FiscalSuggestion.usage_count.desc(), FiscalSuggestion.last_used_at.desc())
            .limit(limit)
        ).all()
    )


def _score_cest_description(product_description: str | None, cest_description: str | None) -> int:
    product_words = {
        word
        for word in normalize_fiscal_description(product_description).split(" ")
        if len(word) >= 4
    }
    cest_words = {
        word
        for word in normalize_fiscal_description(cest_description).split(" ")
        if len(word) >= 4
    }
    if not product_words or not cest_words:
        return 0
    return len(product_words.intersection(cest_words))


def _likely_cest_candidates(
    db: Session,
    *,
    ncm: str,
    product_description: str | None,
    limit: int = 3,
) -> tuple[int, list[MasterFiscalCestCode]]:
    rows = list(
        db.scalars(
            select(MasterFiscalCestCode)
            .where(MasterFiscalCestCode.ncm == ncm)
            .order_by(MasterFiscalCestCode.cest.asc())
        ).all()
    )
    best_by_cest: dict[str, MasterFiscalCestCode] = {}
    for row in rows:
        cest = _digits(row.cest) or row.cest
        current = best_by_cest.get(cest)
        if current is None:
            best_by_cest[cest] = row
            continue
        current_score = _score_cest_description(product_description, current.description)
        row_score = _score_cest_description(product_description, row.description)
        if (row_score, row.updated_at or datetime.min) > (
            current_score,
            current.updated_at or datetime.min,
        ):
            best_by_cest[cest] = row
    ranked = sorted(
        best_by_cest.values(),
        key=lambda row: (
            _score_cest_description(product_description, row.description),
            row.updated_at or datetime.min,
        ),
        reverse=True,
    )
    return len(best_by_cest), ranked[:limit]


def fiscal_alerts_for_product(
    product: Product | None,
    suggestions: list[FiscalSuggestion],
    db: Session | None = None,
) -> list[FiscalAlert]:
    if product is None:
        return []
    alerts: list[FiscalAlert] = []
    if not _clean(product.ncm):
        alerts.append(FiscalAlert(severity="warning", field="ncm", message="Produto sem NCM informado."))
    elif db is not None:
        ncm_code = _digits(product.ncm)
        if not _is_valid_ncm_code(ncm_code):
            alerts.append(FiscalAlert(severity="error", field="ncm", message=f"NCM {product.ncm} invalido: informe um codigo NCM de 8 digitos especifico (nao um capitulo como 1006)."))
            ncm_code = None
        exists = db.scalar(select(MasterFiscalNcmCode).where(MasterFiscalNcmCode.code == ncm_code))
        if ncm_code and exists is None and db.scalar(select(func.count(MasterFiscalNcmCode.id))) > 0:
            alerts.append(FiscalAlert(severity="warning", field="ncm", message=f"NCM {product.ncm} nao encontrado na tabela NCM oficial central."))
    if not _clean(product.cfop_sale):
        alerts.append(FiscalAlert(severity="warning", field="cfop_sale", message="Produto sem CFOP de venda informado."))
    elif db is not None:
        cfop = _digits(product.cfop_sale)
        exists = db.scalar(select(MasterFiscalCfopCode).where(MasterFiscalCfopCode.code == cfop))
        if cfop and exists is None and db.scalar(select(func.count(MasterFiscalCfopCode.id))) > 0:
            alerts.append(FiscalAlert(severity="warning", field="cfop_sale", message=f"CFOP {product.cfop_sale} nao encontrado na tabela CFOP central."))
    if not (_clean(product.cst) or _clean(product.csosn)):
        suggested_tax_status = next(
            (
                ("CSOSN", _clean(suggestion.csosn))
                for suggestion in suggestions
                if _clean(suggestion.csosn)
            ),
            None,
        ) or next(
            (
                ("CST ICMS", _clean(suggestion.cst))
                for suggestion in suggestions
                if _clean(suggestion.cst)
            ),
            None,
        )
        if suggested_tax_status:
            label, value = suggested_tax_status
            alerts.append(
                FiscalAlert(
                    severity="warning",
                    field="cst_csosn",
                    message=(
                        f"Produto sem CST/CSOSN informado. Para venda, pelo regime fiscal da empresa/historico, "
                        f"sugestao: {label} {value}. Conferir com o contador antes de emitir."
                    ),
                )
            )
        else:
            alerts.append(
                FiscalAlert(
                    severity="warning",
                    field="cst_csosn",
                    message=(
                        "Produto sem CST/CSOSN informado. Informe o codigo conforme o regime fiscal da empresa "
                        "e confirme com o contador antes de emitir."
                    ),
                )
            )
    if db is not None and _clean(product.ncm) and not _clean(product.cest):
        cest_count, likely_cests = _likely_cest_candidates(
            db,
            ncm=_digits(product.ncm) or "",
            product_description=product.name,
        )
        if cest_count:
            if likely_cests:
                candidates = "; ".join(
                    f"{row.cest} - {row.description[:80]}" for row in likely_cests
                )
                message = (
                    f"Existem {cest_count} CEST possiveis para o NCM {product.ncm}. "
                    f"Mais provaveis pela descricao: {candidates}. Conferir com o contador."
                )
            else:
                message = (
                    f"Existem {cest_count} CEST possiveis para o NCM {product.ncm} na base central. "
                    "NCM sozinho nao define o CEST; confira descricao, segmento e regra fiscal."
                )
            alerts.append(FiscalAlert(severity="info", field="cest", message=message))
    if _clean(product.cest) is None and any(_clean(suggestion.cest) for suggestion in suggestions):
        alerts.append(FiscalAlert(severity="info", field="cest", message="Existe sugestao com CEST para produto parecido/importado."))
    if not product.new_tax_system and any(
        _clean(suggestion.ibs_cbs_cst)
        or _clean(suggestion.ibs_cbs_classification)
        or _clean(suggestion.selective_tax_cst)
        for suggestion in suggestions
    ):
        alerts.append(FiscalAlert(severity="info", field="new_tax_system", message="XML/importacoes possuem campos IBS/CBS/IS para produto parecido."))
    if db is not None and _clean(product.ibs_cbs_cst) and _clean(product.ibs_cbs_classification):
        reference_count = db.scalar(select(func.count(MasterIbsCbsClassTrib.id))) or 0
        if reference_count:
            reference = db.scalar(
                select(MasterIbsCbsClassTrib).where(
                    MasterIbsCbsClassTrib.cst == _digits(product.ibs_cbs_cst),
                    MasterIbsCbsClassTrib.cclass_trib == _digits(product.ibs_cbs_classification),
                    MasterIbsCbsClassTrib.active.is_(True),
                )
            )
            if reference is None:
                alerts.append(
                    FiscalAlert(
                        severity="warning",
                        field="ibs_cbs_classification",
                        message=(
                            f"Par CST IBS/CBS {product.ibs_cbs_cst} e cClassTrib "
                            f"{product.ibs_cbs_classification} nao encontrado na tabela central."
                        ),
                    )
                )

    compare_fields: tuple[tuple[str, str], ...] = (
        ("ncm", "NCM"),
        ("cfop_sale", "CFOP"),
        ("cst", "CST"),
        ("csosn", "CSOSN"),
        ("ibs_cbs_cst", "CST IBS/CBS"),
        ("ibs_cbs_classification", "cClassTrib IBS/CBS"),
        ("selective_tax_cst", "CST IS"),
    )
    first = suggestions[0] if suggestions else None
    if first is not None:
        for product_field, label in compare_fields:
            suggestion_field = "cfop" if product_field == "cfop_sale" else product_field
            current = _clean(getattr(product, product_field, None))
            suggested = _clean(getattr(first, suggestion_field, None))
            if current and suggested and current != suggested:
                alerts.append(
                    FiscalAlert(
                        severity="info",
                        field=product_field,
                        message=f"{label} do cadastro ({current}) difere da sugestao mais usada/importada ({suggested}).",
                    )
                )
    return alerts
